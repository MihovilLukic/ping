import sys
import subprocess
import threading
import time

from PyQt5.QtGui import QColor
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QApplication, QMainWindow, QTableWidget, QTableWidgetItem, QLineEdit, QLabel, QHBoxLayout, QVBoxLayout, QWidget
from openpyxl import load_workbook

class SpreadsheetApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("IP Addresses Spreadsheet")
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.layout = QVBoxLayout(self.central_widget)
        self.search_layout = QVBoxLayout()
        self.table_widget = QTableWidget()
        self.layout.addLayout(self.search_layout)
        self.layout.addWidget(self.table_widget)
        self.load_data()
        
        from PyQt5.QtCore import QMetaType
        #QMetaType.registerType("QVector<int>")
        
        self.ping_all_thread = threading.Thread(target=self.ping_all)
        self.ping_all_thread.start()
        self.show()

    def load_data(self):
        workbook = load_workbook(filename="data.xlsx")
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        data = []
        for row in sheet.iter_rows(min_row=2):
            data.append([cell.value for cell in row])
        self.table_widget.setColumnCount(len(headers))
        self.table_widget.setHorizontalHeaderLabels(headers)
        self.table_widget.setRowCount(len(data))
        for i, row_data in enumerate(data):
            for j, cell_data in enumerate(row_data):
                item = QTableWidgetItem(str(cell_data))
                item.setFlags(item.flags() ^ Qt.ItemIsEditable)
                self.table_widget.setItem(i, j, item)
        self.table_widget.cellClicked.connect(self.cell_clicked)

        self.search_label = QLabel("Search:")
        self.search_edit = QLineEdit()
        self.search_edit.textChanged.connect(self.filter_table)
        self.search_edit.returnPressed.connect(self.filter_table) # new line
        self.search_layout.addWidget(self.search_label)
        self.search_layout.addWidget(self.search_edit)


    def filter_table(self):
        self.search_edit.returnPressed.connect(self.filter_table)
        search_text = self.search_edit.text()
        for row in range(self.table_widget.rowCount()):
            row_hidden = True
            for column in range(self.table_widget.columnCount()):
                item = self.table_widget.item(row, column)
                if search_text.lower() in item.text().lower():
                    row_hidden = False
                    break
            self.table_widget.setRowHidden(row, row_hidden)

  
    def cell_clicked(self, row, col):
        ip_address = self.table_widget.item(row, col).text()
        if self.ping(ip_address):
            self.table_widget.item(row, col).setForeground(QColor('green'))
        
        else:
            self.table_widget.item(row, col).setForeground(QColor('red'))

    
    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Return or event.key() == Qt.Key_Enter:
            selected_item = self.table_widget.currentItem()
            if selected_item:
                row = selected_item.row()
                col = selected_item.column()
                self.cell_clicked(row, col)
    
    
    def ping(self, ip_address):
        response = subprocess.Popen(["ping", "-n", "1", "-w", "500", ip_address], stdout=subprocess.PIPE).stdout.read()
        return "Reply from" in str(response)

    def ping_all(self):
        while True:
            for row in range(self.table_widget.rowCount()):
                for column in range(self.table_widget.columnCount(3, 4, 5)):
                    item = self.table_widget.item(row, column)
                    ip_address = item.text()
                    if self.ping(ip_address):
                        item.setForeground(QColor('green'))
                    else:
                        item.setForeground(QColor('red'))
            
            time.sleep(1)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    spreadsheet = SpreadsheetApp()
    sys.exit(app.exec_())

