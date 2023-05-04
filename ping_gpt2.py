import sys
import subprocess
import threading
from PyQt5.QtGui import QColor
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QApplication, QMainWindow, QTableWidget, QTableWidgetItem, QLineEdit, QLabel, QHBoxLayout, QVBoxLayout, QWidget, QPlainTextEdit, QPushButton
from openpyxl import load_workbook

class SpreadsheetApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("IP Addresses Spreadsheet")
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.layout = QHBoxLayout(self.central_widget)
        self.table_widget = QTableWidget()
        self.output_widget = QPlainTextEdit()
        self.layout.addWidget(self.table_widget)
        self.layout.addWidget(self.output_widget)
        self.ping_running = False
        self.load_data()
        self.show()

    def load_data(self):
        workbook = load_workbook(filename="data.xlsx")
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        data = []
        for row in sheet.iter_rows(min_row=2):
            data.append([cell.value for cell in row])
        self.table_widget.setColumnCount(len(headers))
        self.table_widget.setHorizontalHeaderLabels(headers)
        self.table_widget.setRowCount(len(data))
        for i, row_data in enumerate(data):
            for j, cell_data in enumerate(row_data):
                item = QTableWidgetItem(str(cell_data))
                item.setFlags(item.flags() ^ Qt.ItemIsEditable)
                self.table_widget.setItem(i, j, item)
        self.table_widget.cellClicked.connect(self.cell_clicked)

        self.search_label = QLabel("Search:")
        self.search_edit = QLineEdit()
        self.search_edit.textChanged.connect(self.filter_table)

        self.ping_button = QPushButton("Ping")
        self.ping_button.setEnabled(False)
        self.ping_button.clicked.connect(self.ping_button_clicked)

        search_layout = QHBoxLayout()
        search_layout.addWidget(self.search_label)
        search_layout.addWidget(self.search_edit)
        search_layout.addWidget(self.ping_button)

        self.layout.addLayout(search_layout)

    def filter_table(self):
        search_text = self.search_edit.text()
        for row in range(self.table_widget.rowCount()):
            row_hidden = True
            for column in range(self.table_widget.columnCount()):
                item = self.table_widget.item(row, column)
                if search_text.lower() in item.text().lower():
                    row_hidden = False
                    break
            self.table_widget.setRowHidden(row, row_hidden)
        
        self.ping_button.setEnabled(False)
        self.output_widget.clear()

    def cell_clicked(self, row, col):
        self.selected_row = row
        self.selected_col = col
        self.ping_button.setEnabled(True)

    def ping_button_clicked(self):
        ip_address = self.table_widget.item(self.selected_row, self.selected_col).text()
        if not self.ping_running:
            self.ping_running = True
            self.ping_button.setText("Stop")
            self.ping_thread = threading.Thread(target=self.ping, args=(ip_address,), daemon=True)
            self.ping_thread.start()
        else:
            self.ping_running = False
            self.ping_button.setText("Ping")
    
    def ping(self, ip_address):
        while self.ping_running:
            process = subprocess.Popen(["ping", "-n", "1", "-w", "500", ip_address], stdout=subprocess.PIPE)
            output, _ = process.communicate()
            if "Reply from" in output.decode():
                self.table_widget.item(self.selected_row, self.selected_col).setForeground(QColor('green'))
            else:
                self.table_widget.item(self.selected_row, self.selected_col)
                
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = SpreadsheetApp()
    sys.exit(app.exec_())                
