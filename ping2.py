import sys
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QApplication, QMainWindow, QTableWidget, QTableWidgetItem
from openpyxl import load_workbook

class SpreadsheetApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("IP Addresses Spreadsheet")
        self.table_widget = QTableWidget()
        self.setCentralWidget(self.table_widget)
        self.load_data()
        self.show()

    def load_data(self):
        workbook = load_workbook(filename="ip_addresses.xlsx")
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        data = []
        for row in sheet.iter_rows(min_row=2):
            data.append([cell.value for cell in row])
        self.table_widget.setColumnCount(len(headers))
        self.table_widget.setHorizontalHeaderLabels(headers)
        self.table_widget.setRowCount(len(data))
        for i, row_data in enumerate(data):
            for j, cell_data in enumerate(row_data):
                item = QTableWidgetItem(str(cell_data))
                item.setFlags(item.flags() ^  Qt.ItemIsEditable)
                self.table_widget.setItem(i, j, item)
        self.table_widget.cellClicked.connect(self.cell_clicked)

    def cell_clicked(self, row, col):
        print(f"Clicked cell {row}, {col}")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    spreadsheet = SpreadsheetApp()
    sys.exit(app.exec_())
